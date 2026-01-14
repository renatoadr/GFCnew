from flask import Blueprint, render_template, redirect, url_for, request, current_app, send_file
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from utils.converter import value_decimal
from decimal import Decimal
from utils.helper import allowed_file
from utils.flash_message import flash_message
from controller.sinapiController import sinapiController
from controller.opcoesController import OpcoesController
from controller.empreendimentoController import empreendimentoController
from dateutil.relativedelta import relativedelta
from utils.security import login_required
from utils.logger import logger
from utils.CtrlSessao import IdEmpreend, NmEmpreend
from datetime import datetime
from zipfile import ZipFile
from io import BytesIO
from enums.indexes_planilha_sinapi import IDX_PLAN_SINAPI
import requests
import openpyxl
import random
import re
import os
from helpers.GerardorPlanilhaSinapi import GerardorPlanilhaSinapi

sinapi_bp = Blueprint('sinapi', __name__)

sinapiCtrl = sinapiController()
opcaoCtrl = OpcoesController()
empreendCtrl = empreendimentoController()

opcaoChave = 'Orca_Sinapi'

url_relatorios = 'https://www.caixa.gov.br/Downloads/sinapi-relatorios-mensais/SINAPI-{vigencia}-formato-xlsx.zip'

user_agents = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.1.1 Safari/605.1.15',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36 OPR/121.0.0.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:143.0) Gecko/20100101 Firefox/143.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36'
]

headers = {
    'User-Agent': '',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Language': 'pt-BR,pt;q=0.9,en;q=0.8',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'cache-control': 'no-cache',
    'accept-language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7',
    'pragma': 'no-cache',
    'priority': 'u=0, i',
    'referer': 'https://www.caixa.gov.br/site/Paginas/downloads.aspx',
    'sec-ch-ua': '"Chromium";v="140", "Not=A?Brand";v="24", "Google Chrome";v="140"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': "Windows",
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'same-origin',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1'
}


@sinapi_bp.route('/atualizar_lista_sinapi', methods=['POST'])
@login_required
def atualizar_lista_sinapi():
    vigencia = (datetime.now() - relativedelta(months=1))
    excel_nome = f"SINAPI_Referência_{vigencia.strftime('%Y_%m')}.xlsx"
    dados = opcaoCtrl.buscar(opcaoChave)

    vigencia = vigencia.replace(day=1)

    if dados is not None and datetime.strptime(dados['mes_referencia'], '%Y-%m-%d') == vigencia:
        flash_message.info("Os dados atuais estão atualizados com a SINAPI")
    else:
        get_sinapi_planilha(vigencia, excel_nome, dados)
    return redirect(url_for('sinapi.sinapi_orcamentos'))


def get_sinapi_planilha(vigencia, excel_nome, dados):
    now = datetime.now()
    try:
        logger.info('Iniciado o carregamento da planilha SINAPI')
        tentativas = Retry(
            total=3,
            status_forcelist=[429, 500, 502, 503, 504, 302],
            allowed_methods=["GET"]
        )
        headers['User-Agent'] = random.choice(user_agents)
        session = requests.Session()
        session.max_redirects = 40
        session.headers.update(headers)
        adaptador = HTTPAdapter(max_retries=tentativas)
        session.mount("http://", adaptador)
        session.mount("https://", adaptador)
        logger.info('Iniciado a solicitação da planilha SINAPI')
        response = session.get(
            url_relatorios.replace(
                '{vigencia}', vigencia.strftime('%Y-%m')),
            timeout=30,
            allow_redirects=True
        )
        logger.info('Finalizada a solicitação da planilha SINAPI')
        response.raise_for_status()
        with ZipFile(BytesIO(response.content)) as zip_file:
            logger.info('Iniciando leitura do zip')
            with zip_file.open(excel_nome) as excel_file:
                logger.info('Iniciando a leitura da planilha')
                wb = openpyxl.load_workbook(
                    filename=excel_file,
                    data_only=False
                )
                logger.info('Carregada planilha para openxl')
                sheetAnalitico = wb['Analítico']
                ref = datetime.strptime(
                    sheetAnalitico['B3'].value, '%m/%Y').date()
                emissao = datetime.strptime(
                    sheetAnalitico['B4'].value, '%d/%m/%Y').date()
                logger.info('Iniciando carregamento do ISD')
                salvarCustoISD(wb['ISD'], now.date(), ref, emissao)
                logger.info('Iniciando carregamento do CSD')
                salvarCustoCSD(wb['CSD'], now.date(), ref, emissao)
                logger.info('Iniciando carregamento dos Itens')
                salvarAnalitico(sheetAnalitico, now.date(), ref, emissao)
                logger.info('Deletando itens antigos')
                deletarAntigos()
                logger.info('Salvando informações de emissão e referência')
                opcaoCtrl.salvar(
                    opcaoChave,
                    {
                        'data_emissao': str(emissao),
                        'mes_referencia': str(ref)
                    }
                )
                logger.info('Finalizado o processamento SINAPI')

    except Exception as err:
        if vigencia > datetime.strptime(dados['mes_referencia'], '%Y-%m-%d'):
            vigencia = (vigencia - relativedelta(months=1))
            excel_nome = f"SINAPI_Referência_{vigencia.strftime('%Y_%m')}.xlsx"
            get_sinapi_planilha(vigencia, excel_nome, dados)
        else:
            logger.exception('Erro ao obter a planilha da SINAPI', err)
            flash_message.error(
                "Não foi possível obter os dados da SINAPI. Tente novamente mais tarde")


@sinapi_bp.route('/download_orcamento_sinapi/<nome_arquivo>')
@login_required
def download_orcamento_sinapi(nome_arquivo):
    file = os.path.join(
        current_app.config['DIRSYS'],
        "orcamentos_sinapi",
        IdEmpreend().get(),
        nome_arquivo
    )
    return send_file(file)


@sinapi_bp.route('/deletar_orcamento_sinapi/<nome_arquivo>')
@login_required
def deletar_orcamento_sinapi(nome_arquivo):
    file = os.path.join(
        current_app.config['DIRSYS'],
        "orcamentos_sinapi",
        IdEmpreend().get(),
        nome_arquivo
    )
    if os.path.exists(file):
        os.remove(file)
    return redirect('/sinapi_orcamentos')


@sinapi_bp.route('/sinapi_orcamentos')
@login_required
def sinapi_orcamentos():
    idEmpreend = request.args.get("idEmpreend")
    nmEmpreend = request.args.get('nmEmpreend')
    dados = opcaoCtrl.buscar(opcaoChave)

    if (idEmpreend is None and not IdEmpreend().has()) or (nmEmpreend is None and not NmEmpreend().has()):
        return redirect('/home')

    if idEmpreend is None:
        idEmpreend = IdEmpreend().get()
    else:
        IdEmpreend().set(idEmpreend)

    if nmEmpreend is None:
        nmEmpreend = NmEmpreend().get()
    else:
        NmEmpreend().set(nmEmpreend)

    path_save = os.path.join(
        current_app.config['DIRSYS'], "orcamentos_sinapi", IdEmpreend().get())
    orcs = []

    last_date = None
    if os.path.exists(path_save):
        for orc in os.listdir(path_save):
            infos = orc.split('_')
            current_date = datetime.strptime(infos[-2], '%Y-%m-%d')
            obj = {
                'criacao': converterData(infos[-1].replace('.xlsx', '')),
                'referencia': converterData(infos[-2], False),
                'nome_arquivo': orc
            }
            if last_date is None or current_date > last_date:
                orcs.append(obj)
            else:
                orcs.insert(0, obj)
            last_date = current_date
    return render_template(
        "lista_orcamentos_sinapi.html",
        emissao='--' if dados is None else converterData(
            dados['data_emissao']),
        ref='--' if dados is None else converterData(
            dados['mes_referencia'], False),
        orcs=orcs
    )


@sinapi_bp.route('/upload_orcamento_sinapi', methods=['POST'])
@login_required
def upload_orcamento_sinapi():
    dados = opcaoCtrl.buscar(opcaoChave)
    if 'file' not in request.files:
        flash_message.error(
            'Não existe uma planilha para realizar o processamento')
        return redirect(url_for('sinapi.sinapi_orcamentos'))

    file = request.files['file']
    if not allowed_file(file.filename):
        flash_message.error(
            'O arquivo informado não é uma planiha e não pode ser processado')
    elif dados is None or not sinapiController.existe_itens():
        flash_message.error('Não existem dados para processar o documento')
    else:
        processarPlanilha(file)
        flash_message.success(
            'Processamento concluído! Faça o download do arquivo na lista abaixo.')
    return redirect(url_for('sinapi.sinapi_orcamentos'))


@sinapi_bp.route('/gerar_planilha_base_sinapi')
@login_required
def gerar_planilha_base_sinapi():
    try:
        prods = sinapiController.buscar_categorias_produtos()
        file = os.path.join(
            __file__.replace('sinapi.py', ''),
            '..',
            'static',
            'planilha_base_orcamento_sinapi.xlsx'
        )
        gerador = GerardorPlanilhaSinapi(file, prods)
        gerador.gerar_planilha_base_sinapi()
        return send_file(
            gerador.get_stream(),
            as_attachment=True,
            download_name='planilha_base_orcamento_sinapi.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.exception('Erro ao gerar planilha base de orçamento SINAPI', e)
        flash_message.error('Erro ao gerar planilha base de orçamento SINAPI')
        return redirect(url_for('sinapi.sinapi_orcamentos'))


def processarPlanilha(file):
    dados = opcaoCtrl.buscar(opcaoChave)
    empreend = empreendCtrl.consultarEmpreendimentoPeloId(
        IdEmpreend().get())
    path_save = os.path.join(
        current_app.config['DIRSYS'], "orcamentos_sinapi", IdEmpreend().get())

    if not os.path.exists(path_save):
        os.makedirs(path_save)

    ws = openpyxl.load_workbook(
        filename=file,
        data_only=True
    )
    sheet = ws.active
    sheet[IDX_PLAN_SINAPI.NomeCliente.value] = empreend.getNmIncorp()
    sheet[IDX_PLAN_SINAPI.Endereco.value] = empreend.getEnderecoCompleto()
    sheet[IDX_PLAN_SINAPI.Data.value] = datetime.now().strftime('%d/%m/%Y')
    sheet[IDX_PLAN_SINAPI.NomeEmpreend.value] = empreend.getNmEmpreend()
    sheet[IDX_PLAN_SINAPI.DescPercentDiff.value] = 'Diferença entre orçamentos'
    sheet[IDX_PLAN_SINAPI.DescTotalSinapi.value] = 'Orçamento SINAPI (França Part.)'
    sheet[IDX_PLAN_SINAPI.CabecalhoTotal.value] = 'Total Obra'
    sheet[IDX_PLAN_SINAPI.CabecalhoUnitSinapi.value] = 'Custo unit Sinapi (França Part.)'
    sheet[IDX_PLAN_SINAPI.CabecalhoTotalSinapi.value] = 'Total Sinapi (França Part.)'
    sheet[IDX_PLAN_SINAPI.CabecalhoPercentDiff.value] = 'Diferença Orçamento'

    total = 0
    totalSinapi = 0

    for row in sheet.iter_rows(min_row=9):
        codigo = row[IDX_PLAN_SINAPI.Codigo.value].value
        quantidade = row[IDX_PLAN_SINAPI.Quantidade.value].value
        quantEng = row[IDX_PLAN_SINAPI.QuantEng.value].value
        custoUnit = row[IDX_PLAN_SINAPI.Custo.value].value

        if quantidade is None and custoUnit is None and quantEng is None:
            continue

        if quantidade is not None and (isinstance(quantidade, int) or isinstance(quantidade, float)):
            quantidade = Decimal(quantidade)
        else:
            quantidade = Decimal(0)

        if quantEng is not None and (isinstance(quantEng, int) or isinstance(quantEng, float)):
            quantEng = Decimal(quantEng)
        else:
            quantEng = Decimal(0)

        if custoUnit is not None and (isinstance(custoUnit, int) or isinstance(custoUnit, float)):
            custoUnit = Decimal(custoUnit)
        else:
            custoUnit = Decimal(0)

        if quantEng == 0:
            continue

        totalManual = custoUnit * quantidade
        row[IDX_PLAN_SINAPI.Total.value].value = totalManual
        total += totalManual

        totalSinapiLinha = Decimal(0)
        if codigo is not None and isinstance(codigo, int):
            contadores = sinapiController.valor_total_itens(
                IdEmpreend().get(),
                codigo,
                quantEng
            )
            if contadores['total'] == 0:
                totalEng = custoUnit * quantEng
                row[IDX_PLAN_SINAPI.UnitSinapi.value].value = custoUnit
                row[IDX_PLAN_SINAPI.TotalSinapi.value].value = totalEng
                totalSinapiLinha = totalEng
                totalSinapi += totalEng
            else:
                row[IDX_PLAN_SINAPI.UnitSinapi.value].value = contadores['vl_unitario']
                row[IDX_PLAN_SINAPI.TotalSinapi.value].value = contadores['total']
                totalSinapi += contadores['total']
                totalSinapiLinha = contadores['total']
        else:
            totalEng = custoUnit * quantEng
            row[IDX_PLAN_SINAPI.UnitSinapi.value].value = custoUnit
            row[IDX_PLAN_SINAPI.TotalSinapi.value].value = totalEng
            totalSinapiLinha = totalEng
            totalSinapi += totalEng

        percentdiff = calcPercent(totalManual, totalSinapiLinha)
        if codigo is None and custoUnit == 0:
            row[IDX_PLAN_SINAPI.PercentDiff.value].value = 0
        elif quantidade == 0:
            row[IDX_PLAN_SINAPI.PercentDiff.value].value = 1
        elif totalManual == totalSinapiLinha:
            row[IDX_PLAN_SINAPI.PercentDiff.value].value = 0
        elif totalManual > totalSinapiLinha:
            row[IDX_PLAN_SINAPI.PercentDiff.value].value = percentdiff * (-1)
        else:
            row[IDX_PLAN_SINAPI.PercentDiff.value].value = percentdiff

    percentdiffTotal = calcPercent(total, totalSinapi)
    sheet[IDX_PLAN_SINAPI.ValorTotal.value] = value_decimal(total)
    sheet[IDX_PLAN_SINAPI.ValorTotalSinapi.value] = value_decimal(totalSinapi)
    if total == totalSinapi:
        sheet[IDX_PLAN_SINAPI.ValorPercentDiff.value] = 0
    elif total > totalSinapi:
        sheet[IDX_PLAN_SINAPI.ValorPercentDiff.value] = percentdiffTotal * (-1)
    else:
        sheet[IDX_PLAN_SINAPI.ValorPercentDiff.value] = percentdiffTotal

    ws.save(filename=os.path.join(
        path_save, f'orcamento_sinapi_{NmEmpreend().get().replace(' ', '_')}_{dados['mes_referencia']}_{datetime.now().date()}.xlsx'))


def calcPercent(valor1, valor2):
    if valor1 == 0 or valor2 == 0:
        return 0
    maior = max(valor1, valor2)
    menor = min(valor1, valor2)
    return (maior - menor) / maior


def salvarCustoISD(isds, dt_corrente, dt_vigencia, dt_emissao):
    itens_salvar = []

    for row in isds.iter_rows(min_row=11):
        itens_salvar.append((
            str(dt_corrente),
            str(dt_vigencia),
            str(dt_emissao),
            row[1].value,
            'ISD',
            0.0 if row[5].value is None else row[5].value,
            0.0 if row[6].value is None else row[6].value,
            0.0 if row[7].value is None else row[7].value,
            0.0 if row[8].value is None else row[8].value,
            0.0 if row[9].value is None else row[9].value,
            0.0 if row[10].value is None else row[10].value,
            0.0 if row[11].value is None else row[11].value,
            0.0 if row[12].value is None else row[12].value,
            0.0 if row[13].value is None else row[13].value,
            0.0 if row[14].value is None else row[14].value,
            0.0 if row[15].value is None else row[15].value,
            0.0 if row[16].value is None else row[16].value,
            0.0 if row[17].value is None else row[17].value,
            0.0 if row[18].value is None else row[18].value,
            0.0 if row[19].value is None else row[19].value,
            0.0 if row[20].value is None else row[20].value,
            0.0 if row[21].value is None else row[21].value,
            0.0 if row[22].value is None else row[22].value,
            0.0 if row[23].value is None else row[23].value,
            0.0 if row[24].value is None else row[24].value,
            0.0 if row[25].value is None else row[25].value,
            0.0 if row[26].value is None else row[26].value,
            0.0 if row[27].value is None else row[27].value,
            0.0 if row[28].value is None else row[28].value,
            0.0 if row[29].value is None else row[29].value,
            0.0 if row[30].value is None else row[30].value,
            0.0 if row[31].value is None else row[31].value,
        ))
    sinapiCtrl.inserir_custos(itens_salvar)


def salvarCustoCSD(csds, dt_corrente, dt_vigencia, dt_emissao):
    itens_salvar = []

    for row in csds.iter_rows(min_row=11):
        itens_salvar.append((
            str(dt_corrente),
            str(dt_vigencia),
            str(dt_emissao),
            int(re.findall(r"\d{3,}", row[1].value)[0]),
            'CSD',
            0.0 if row[4].value is None else row[4].value,
            0.0 if row[6].value is None else row[6].value,
            0.0 if row[8].value is None else row[8].value,
            0.0 if row[10].value is None else row[10].value,
            0.0 if row[12].value is None else row[12].value,
            0.0 if row[14].value is None else row[14].value,
            0.0 if row[16].value is None else row[16].value,
            0.0 if row[18].value is None else row[18].value,
            0.0 if row[20].value is None else row[20].value,
            0.0 if row[22].value is None else row[22].value,
            0.0 if row[24].value is None else row[24].value,
            0.0 if row[26].value is None else row[26].value,
            0.0 if row[28].value is None else row[28].value,
            0.0 if row[30].value is None else row[30].value,
            0.0 if row[32].value is None else row[32].value,
            0.0 if row[34].value is None else row[34].value,
            0.0 if row[36].value is None else row[36].value,
            0.0 if row[38].value is None else row[38].value,
            0.0 if row[40].value is None else row[40].value,
            0.0 if row[42].value is None else row[42].value,
            0.0 if row[44].value is None else row[44].value,
            0.0 if row[46].value is None else row[46].value,
            0.0 if row[48].value is None else row[48].value,
            0.0 if row[50].value is None else row[50].value,
            0.0 if row[52].value is None else row[52].value,
            0.0 if row[54].value is None else row[54].value,
            0.0 if row[56].value is None else row[56].value,
        ))
    sinapiCtrl.inserir_custos(itens_salvar)


def salvarAnalitico(analiticos, dt_corrente, dt_vigencia, dt_emissao):
    itens_salvar = []

    for row in analiticos.iter_rows(min_row=11):
        itens_salvar.append((
            str(dt_corrente),
            str(dt_vigencia),
            str(dt_emissao),
            row[3].value,
            row[1].value,
            row[0].value,
            row[2].value,
            row[4].value,
            row[5].value,
            0.0 if row[6].value is None else float(row[6].value),
        ))

    sinapiCtrl.inserir_itens(itens_salvar)


def deletarAntigos():
    data_ref = (datetime.now() - relativedelta(months=3)).strftime('%Y-%m-%d')
    sinapiController.deletar_custos(data_ref)
    sinapiController.deletar_itens(data_ref)


def converterData(data, usarDia=True):
    if usarDia:
        return datetime.strptime(data, '%Y-%m-%d').strftime('%d/%m/%Y')
    else:
        return datetime.strptime(data, '%Y-%m-%d').strftime('%m/%Y')
