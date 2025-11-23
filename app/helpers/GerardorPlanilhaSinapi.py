from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Protection, PatternFill, Side, Font, Border, Alignment

from dto.produto import produto


class GerardorPlanilhaSinapi:
    lmt_omissos = 5
    init_row = 7

    ws = None
    sheet = None
    current_prod = None
    next_prod = None
    old_prod = None
    is_last = False
    row_init_group = init_row
    row_start = init_row
    total_prods = 0
    pends = 0

    def __init__(self, file: str, prods: list[produto]):
        self.ws = load_workbook(filename=file)
        self.sheet = self.ws.active
        self.prods = prods

    def border_config(self):
        return Side(
            border_style="thin",
            color="D9D9D9"
        )

    def font(self, color='000000', size=11, bold=False):
        return Font(
            bold=bold,
            size=size,
            color=color,
            name='Consolas'
        )

    def borders(self):
        return Border(
            top=self.border_config(),
            left=self.border_config(),
            right=self.border_config(),
            bottom=self.border_config()
        )

    def alignments(self, hcenter=True):
        return Alignment(
            horizontal='center' if hcenter else 'left',
            vertical='center',
            wrap_text=True
        )

    def fill(self, color='F2F2F2'):
        return PatternFill(
            start_color=color,
            end_color=color,
            fill_type='solid'
        )

    def merge(self, start: int, end: int):
        self.sheet.merge_cells(
            start_row=self.row_start,
            end_row=self.row_start,
            start_column=start,
            end_column=end
        )

    def valor_cel(self, col: int, valor):
        return self.sheet.cell(self.row_start, col, valor)

    def altura_linha(self, altura=15):
        self.sheet.row_dimensions[self.row_start].height = altura

    def inc_row(self):
        self.row_start += 1

    def cell_basic_config(self):
        col = 1
        while col <= 12:
            cell = self.sheet.cell(self.row_start, col)
            cell.border = self.borders()
            cell.protection = Protection(locked=False)
            cell.alignment = self.alignments()
            cell.font = self.font('111111', 9)
            col += 1

    def config_categoria(self):
        if self.old_prod is None or self.current_prod.getCategoriaPai() != self.old_prod.getCategoriaPai():
            cell = self.valor_cel(1, self.current_prod.getCategoriaPai())
            cell.alignment = self.alignments()
            cell.fill = self.fill('A6A6A6')
            cell.font = self.font('262626', 10, True)
            cell.border = self.borders()
            cell.protection = Protection(locked=True)
            self.altura_linha()
            self.merge(1, 12)
            self.inc_row()

    def config_subcategoria(self):
        if self.old_prod is None or self.current_prod.getCategoria() != self.old_prod.getCategoria():
            cell = self.valor_cel(1, self.current_prod.getCategoria())
            cell.fill = self.fill('F2F2F2')
            cell.font = self.font('31869B', 9, True)
            cell.alignment = self.alignments(hcenter=False)
            cell.border = self.borders()
            cell.protection = Protection(locked=True)
            self.altura_linha(20)
            self.merge(1, 12)
            self.inc_row()

    def config_omissos(self):
        if self.next_prod is None or self.current_prod.getCategoriaPai() != self.next_prod.getCategoriaPai():
            cell = self.valor_cel(1, 'Omissos')
            cell.alignment = self.alignments()
            cell.fill = self.fill('F2F2F2')
            cell.font = self.font('808080', 9, True)
            cell.border = self.borders()
            cell.protection = Protection(locked=True)
            self.altura_linha(16)
            self.merge(1, 12)
            self.inc_row()

            jdx = 0
            while jdx < self.lmt_omissos:
                self.merge(2, 4)
                self.cell_basic_config()
                self.inc_row()
                jdx += 1

    def config_produto(self):
        self.cell_basic_config()
        desc = self.current_prod.getDescricao()
        code = self.current_prod.getCodigo()

        cell = self.valor_cel(1, code)
        cell.alignment = self.alignments()
        cell.font = self.font('D9D9D9', 9)
        cell.protection = Protection(locked=True if code else False)

        cell = self.valor_cel(2, self.current_prod.getDescricao())
        cell.alignment = self.alignments(hcenter=False)
        cell.font = self.font('111111', 9)
        cell.protection = Protection(locked=True)

        cell = self.valor_cel(5, self.current_prod.getUnidade())
        cell.alignment = self.alignments()
        cell.font = self.font('111111', 9)
        cell.protection = Protection(locked=True)

        if len(desc) > 180:
            self.altura_linha(46)
        elif len(desc) > 95:
            self.altura_linha(34)
        else:
            self.altura_linha(20)

        self.merge(2, 4)
        self.inc_row()

    def config_topics(self):
        if self.is_last or (self.current_prod is not None and self.old_prod is not None and self.current_prod.getCategoriaPai() != self.old_prod.getCategoriaPai() and self.row_start - 1 > self.row_init_group):
            init = int(self.row_init_group)
            end = int(self.row_start) - \
                2 if not self.is_last else self.pends + self.init_row + self.row_start
            self.sheet.row_dimensions.group(init, end)
            self.row_init_group = self.row_start

    def set_format_data(self, col: int, format: str):
        for row in self.sheet.iter_rows(min_row=self.init_row, min_col=col, max_col=col):
            for cell in row:
                cell.number_format = format

    def gerar_planilha_base_sinapi(self):
        self.total_prods = len(self.prods)
        self.pends = self.total_prods

        for idx in range(0, self.total_prods):
            self.is_last = idx == self.total_prods - 1
            self.current_prod = self.prods[idx]

            if idx > 0:
                self.old_prod = self.prods[idx - 1]

            if idx < self.total_prods - 1:
                self.next_prod = self.prods[idx + 1]
            else:
                self.next_prod = None

            self.config_categoria()
            self.config_topics()
            self.config_subcategoria()
            self.config_produto()
            self.config_omissos()
            self.pends -= 1

        self.set_format_data(6, '#,##0.00')
        self.set_format_data(7, 'R$ #,##0.00')
        self.set_format_data(8, '#,##0.00')
        self.set_format_data(9, 'R$ #,##0.00')
        self.set_format_data(10, 'R$ #,##0.00')
        self.set_format_data(11, 'R$ #,##0.00')
        self.set_format_data(12, '0.00%')

        self.sheet.protection.sheet = True
        self.sheet.protection.set_password('GfC@4E96')

    def get_stream(self):
        file_stream = BytesIO()
        self.ws.save(file_stream)
        file_stream.seek(0)
        return file_stream
