import datetime
from openpyxl import Workbook


class DebugRelatorioUnidade:
    unidades: list = []
    unidadesProcessadas: list = []
    totais: dict = {}
    totaisProcessados: dict = {}

    def __init__(self, idEmpreend: int, dataInicio: datetime, dataFim: datetime):
        self.idEmpreend = idEmpreend
        self.dataInicio = dataInicio
        self.dataFim = dataFim

    def __sheetUnidades__(self, wb: Workbook):
        sheet = wb.create_sheet("Unidades")
        sheet.append(["id_unidade", "id_empreendimento", "id_torre", "id_empreed_elonet", "id_torre_elonet", "id_unidade_elonet", "unidade", "mes_vigencia", "ano_vigencia",
                      "vl_unidade", "status", "cpf_cnpj_comprador", "vl_receber", "dt_ocorrencia", "financiado", "vl_chaves", "vl_pre_chaves", "vl_pos_chaves", "ac_historico"])
        for unidade in self.unidades:
            sheet.append([
                unidade.get("id_unidade"),
                unidade.get("id_empreendimento"),
                unidade.get("id_torre"),
                unidade.get("id_empreed_elonet"),
                unidade.get("id_torre_elonet"),
                unidade.get("id_unidade_elonet"),
                unidade.get("unidade"),
                unidade.get("mes_vigencia"),
                unidade.get("ano_vigencia"),
                unidade.get("vl_unidade"),
                unidade.get("status"),
                unidade.get("cpf_cnpj_comprador"),
                unidade.get("vl_receber"),
                unidade.get("dt_ocorrencia"),
                unidade.get("financiado"),
                unidade.get("vl_chaves"),
                unidade.get("vl_pre_chaves"),
                unidade.get("vl_pos_chaves"),
                unidade.get("ac_historico"),
            ])

    def __sheetUnidadesProcessadas__(self, wb: Workbook):
        sheet = wb.create_sheet("Unidades Processadas")
        sheet.append(["id_unidade", "id_torre", "id_empreendimento", "unidade", "mes_vigencia", "ano_vigencia", "vl_unidade", "status",
                     "cpf_cnpj_comprador", "vl_receber", "dt_ocorrencia", "financiado", "vl_chaves", "vl_pre_chaves", "vl_pos_chaves", "nm_torre"])
        for uni in self.unidadesProcessadas:
            sheet.append([
                uni.getIdUnidade(),
                uni.getIdTorre(),
                uni.getIdEmpreend(),
                uni.getUnidade(),
                uni.getMesVigencia(),
                uni.getAnoVigencia(),
                uni.getVlUnidade(),
                uni.getStatus(),
                uni.getCpfComprador(),
                uni.getVlReceber(),
                uni.getDtOcorrencia(),
                uni.getFinanciado(),
                uni.getVlChaves(),
                uni.getVlPreChaves(),
                uni.getVlPosChaves(),
                uni.getNmTorre(),
            ])

    def __sheetTotais__(self, wb: Workbook):
        sheet = wb.create_sheet("Totais")
        sheet.append(["Vigência", "Valor Unidade",
                     "Valor Pago",  "Valor Receber", ])
        for key, value in self.totais.items():
            sheet.append([
                key,
                value["valorUnidade"],
                value["valorPago"],
                value["valorReceber"],
            ])

    def __sheetTotaisProcessados__(self, wb: Workbook):
        sheet = wb.create_sheet("Totais Processados")
        sheet.append(["Mês Vigência", "Ano Vigência",
                     "Total Unidade", "Total Pago"])
        for uni in self.totaisProcessados:
            sheet.append([
                uni.getMesVigencia(),
                uni.getAnoVigencia(),
                uni.getTtUnidade(),
                uni.getTtPago(),
            ])

    def save(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Dados Entrada"
        sheet["A1"] = "ID Empreendimento"
        sheet["A2"] = "Data Início"
        sheet["A3"] = "Data Fim"
        sheet["B1"] = self.idEmpreend
        sheet["B2"] = self.dataInicio.strftime("%Y-%m-%d")
        sheet["B3"] = self.dataFim.strftime("%Y-%m-%d")
        self.__sheetUnidades__(wb)
        self.__sheetUnidadesProcessadas__(wb)
        self.__sheetTotais__(wb)
        self.__sheetTotaisProcessados__(wb)
        filename = f"Debug_Relatorio_Unidade_{self.idEmpreend}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        wb.save(filename)
